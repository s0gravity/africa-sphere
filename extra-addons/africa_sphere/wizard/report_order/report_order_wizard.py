# -*- encoding: utf-8 -*-

from openerp import models, fields, api, exceptions, _
import datetime
import os
from cStringIO import StringIO
import base64
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from dateutil import parser

class report_order_wizard(models.TransientModel):
    _name = 'report.order.wizard'

    partner_id = fields.Many2one(comodel_name="res.partner", string="Client", domain=[('is_company','=',True),('customer','=',True)])
    product_id = fields.Many2one(comodel_name="product.product", string="Article", domain=[('is_report','=',True)])
    action = fields.Selection(string="Action", selection=[('new', 'Nouveau BC'), ('merge', 'Fusionner'), ], default='new')
    sale_order_id = fields.Many2one(comodel_name="sale.order", string="Bon de commande", domain=[('state','=','draft')])
    company_fiscalyear_id1 = fields.Many2one(comodel_name="account.fiscalyear", string="Exercice N")
    company_fiscalyear_id2 = fields.Many2one(comodel_name="account.fiscalyear", string="Exercice N-1")

    #informations sur commande
    document_to = fields.Char(string="DOCUMENT A L'USAGE EXCLUSIF DE")
    partners_ref = fields.Char(string="Références Clients / Ref. De la commande")
    request_date = fields.Date(string="Date de la requête")
    response_date = fields.Date(string="Date de réponse de @SPHERES")
    response_delay = fields.Integer(string="Delais de Réponse (en jours)",compute="_get_response_delay")
    customer_internal_code = fields.Char(string="CODE CLIENT INTERNE")
    credit_solicite = fields.Float(string="ENCOURS CREDIT SOLICITE",digits=(6,2))
    credit_rec = fields.Float(string="ENCOURS CREDIT RECOMMANDE PAR @SPHERES",digits=(6,2))
    credit_approval_percent = fields.Float(string="% Accord Crédit @SPHERES",digits=(6,2),compute="_credit_approval_percent")
    credit_currency_id = fields.Many2one(comodel_name="res.currency", string="Devise ENCOURS")
    comment_sphere = fields.Text(string="COMMENTAIRES")
    principal_partner_ids = fields.One2many(comodel_name="principal.partner", inverse_name="sale_order_id", string="PARTENAIRES PRINCIPAUX")
    strenght1 = fields.Text(string="1")
    strenght2 = fields.Text(string="2")
    strenght3 = fields.Text(string="3")
    strenght4 = fields.Text(string="4")
    strenght5 = fields.Text(string="5")
    weakness1 = fields.Text(string="1")
    weakness2 = fields.Text(string="2")
    weakness3 = fields.Text(string="3")
    weakness4 = fields.Text(string="4")
    weakness5 = fields.Text(string="5")
    short_term_dyn = fields.Text(string="Dynamisme à Court Terme")
    short_term_risks = fields.Text(string="Risques Potentiels à Court Terme")
    company_fiscalyear_id1 = fields.Many2one(comodel_name="account.fiscalyear", string="Exercice N")
    company_fiscalyear_id2 = fields.Many2one(comodel_name="account.fiscalyear", string="Exercice N-1")

    @api.one
    @api.depends('request_date','response_date')
    def _get_response_delay(self):
        if self.request_date and self.response_date:
            self.response_delay = ((parser.parse(self.response_date).date()) - (parser.parse(self.request_date).date())).days

    @api.one
    @api.depends('credit_solicite','credit_rec')
    def _credit_approval_percent(self):
        if self.credit_solicite and self.credit_rec:
            self.credit_approval_percent = self.credit_rec/self.credit_solicite

    #informations sur commande

    #Couleurs ratios
    ratio_fond_roulement = fields.Selection(string="FOND DE ROULEMENT NET = FDR", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_rotation_stock = fields.Selection(string="ROTATION DES STOCKS en jours", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_besoin_fond_roulement = fields.Selection(string="BESOIN EN FOND DE ROULEMENT = BFR", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_clients = fields.Selection(string="ROTATION DES CLIENTS en jours", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_treso_net = fields.Selection(string="TRESORERIE NETTE", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_rotation_fournisseurs = fields.Selection(string="ROTATION DES FOURNISSEURS en jours", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_liquid_test = fields.Selection(string="LIQUIDITE 'Liquid Test' EN % (> 1)", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_degree_endettement = fields.Selection(string="DEGRE ENDETTEMENT EN %", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_autonomie_financiere = fields.Selection(string="DEGRE AUTONOMIE FINANCIERE (= ou > 20%)", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_dettes_fin_cafg = fields.Selection(string="DETTES FIN. / CAFG", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_rn_caff_ht = fields.Selection(string="RN / CAFF HT % (> 0%)", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_cont_exp = fields.Selection(string="CONTINUE EXPLOITATION EN % (> 50%)", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    ratio_capacite_fin = fields.Selection(string="Capacité d'autofinancement ou CAFG", selection=[('green', 'Vert'), ('orange', 'Orange'),('red','Rouge') ])
    #Couleurs ratios


    def _print_excel(self,report_name,model,res_id):

        if report_name == "africa_sphere.product_A_template":
            buf=StringIO()
            #############formatage des cellules###############################"
            ft= Font(name='Calibri',size=14,bold=True,italic=True,vertAlign=None,underline='none',strike=False,color='FF000000')
            ft2= Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
            fl=PatternFill(fill_type='solid',start_color='FF2AF21C',end_color='FF2AF21C')
            ###############################definition du fichier excel##############
            wb = Workbook(guess_types=True)
            ws = wb.active
            ws.title = "Rapport"
            ####################################remplir la feuille excel avec les données########################################
            title1 = unicode('AFRICA SHPERE (PRODUCT A)', "utf8")
            ce=ws.cell(column=4, row=2, value=title1)
            ce.font=ft
            header = ['Country Phone code','Country name','CENTRAL BANK Activity CODE','CENTRAL BANK Activity MAIN NAME','Activity wording','COMPANY NAME',
                      'REGISTRATION NUMBER','CAPITAL = SHARES','TURNOVER YEAR-1','TURNOVER YEAR-2','NET RESULT YEAR-1','NET RESULT YEAR-2','@SPHERES ONGOING SCORING']
            header_mapping = ['country_phone_code','country_id.name','central_bank_activity_code','l_bank_activity_main','activity_wording','name',
                              'registration_number','capital','turnover_y1','turnover_y2','net_result_y1','net_result_y2','scoring']
            i = 1
            for val in header:
                ce=ws.cell(column=i, row=4, value=val)
                ce.font=ft
                ce.fill=fl
                i+=1

            j= 5
            companies = self.env['sphere.company'].search([('id','in',self._context['active_ids'])])
            for c in companies:
                i = 1
                for fld in header_mapping:
                    exec "ce=ws.cell(column=i, row=j, value=c."+fld+")"
                    ce.font=ft2
                    i+=1
                j+=1


            ###################################enrgistrement du fichier####################################
            wb.save(buf)
            fichier = self.product_id.name+".xlsx"
            out=base64.encodestring(buf.getvalue())
            buf.close()
            ###################################Pièce jointe####################################
            ir_attachment = self.env['ir.attachment'].create(
                                                              {'name': fichier,
                                                               'datas': out,
                                                               'datas_fname': fichier,
                                                               'res_model': model,
                                                               'res_id':res_id
                                                               })


        return True


    @api.multi
    def apply_pricelist(self, pricelist, product_id, qty, amount, date):
        product = self.env['product.product'].browse(product_id)
        if not pricelist:
            return amount
        else:
            ctx = dict(
                uom=product.uom_id.id,
                date=date,
            )
            #taxes
            account = product.property_account_income or product.categ_id.property_account_income_categ
            taxes = product.taxes_id or account.tax_ids
            fpos = self.env['account.fiscal.position'].browse(False)
            fp_taxes = fpos.map_tax(taxes)
            #taxes
            price = self.pool.get('product.pricelist').price_get(self._cr, self._uid, [pricelist],
                    product.id, qty or 1.0, self.partner_id.id, ctx)[pricelist]
            if price is False:
                return amount
            else:
                price = self.pool['account.tax']._fix_tax_included_price(self._cr, self._uid, price, taxes, fp_taxes.ids)
                amount = price
                #print "price unit 3 =",price_unit
                return amount

    @api.multi
    def action_make_order(self):

        if self.action == 'new':

            sale_obj = self.env['sale.order']
            payment_term = self.partner_id.property_payment_term and self.partner_id.property_payment_term.id or False
            partner_addr = self.pool.get('res.partner').address_get(self._cr, self._uid, [self.partner_id.id],['default', 'invoice', 'delivery', 'contact'])
            pricelist = self.partner_id.property_product_pricelist.id
            fpos = self.partner_id.property_account_position and self.partner_id.property_account_position.id or False
            sale_order_vals = {
                    'origin': "Report",
                    'partner_id': self.partner_id.id,
                    'pricelist_id': pricelist,
                    'partner_invoice_id': partner_addr['invoice'],
                    'partner_shipping_id': partner_addr['delivery'],
                    'date_order': fields.datetime.now(),
                    'fiscal_position': fpos,
                    'payment_term':payment_term,
                    }

            sale_order = sale_obj.create(sale_order_vals)


            price_unit = self.apply_pricelist(pricelist, self.product_id.id, 1.00, self.product_id.list_price, sale_order.date_order)
            sale_order_line_vals ={
                'order_id':sale_order.id,
                'product_id':self.product_id.id,
                'name':self.product_id.name,
                'product_uom':self.product_id.uom_id.id,
                'product_uom_qty':1,
                'price_unit':price_unit
            }

            new_line = self.env['sale.order.line'].create(sale_order_line_vals)
            new_line.tax_id = self.product_id.taxes_id


        else:
            sale_order = self.sale_order_id
            price_unit = self.apply_pricelist(sale_order.pricelist_id.id, self.product_id.id, 1.00, self.product_id.list_price, sale_order.date_order)
            sale_order_line_vals ={
                'order_id':sale_order.id,
                'product_id':self.product_id.id,
                'name':self.product_id.name,
                'product_uom':self.product_id.uom_id.id,
                'product_uom_qty':1,
                'price_unit':price_unit
            }
            new_line = self.env['sale.order.line'].create(sale_order_line_vals)
            new_line.tax_id = self.product_id.taxes_id

        #informations sur commande
        sale_order.document_to = self.document_to
        sale_order.partners_ref = self.partners_ref
        sale_order.request_date = self.request_date
        sale_order.response_date = self.response_date
        sale_order.response_delay = self.response_delay
        sale_order.customer_internal_code = self.customer_internal_code
        sale_order.credit_solicite = self.credit_solicite
        sale_order.credit_rec = self.credit_rec
        sale_order.credit_approval_percent = self.credit_approval_percent
        sale_order.comment_sphere = self.comment_sphere
        sale_order.principal_partner_ids = self.principal_partner_ids
        sale_order.strenght1 = self.strenght1
        sale_order.strenght2 = self.strenght2
        sale_order.strenght3 = self.strenght3
        sale_order.strenght4 = self.strenght4
        sale_order.strenght5 = self.strenght5
        sale_order.weakness1 = self.weakness1
        sale_order.weakness2 = self.weakness2
        sale_order.weakness3 = self.weakness3
        sale_order.weakness4 = self.weakness4
        sale_order.weakness5 = self.weakness5
        sale_order.short_term_dyn = self.short_term_dyn
        sale_order.short_term_risks = self.short_term_risks
        sale_order.company_fiscalyear_id1 = self.company_fiscalyear_id1
        sale_order.company_fiscalyear_id2 = self.company_fiscalyear_id2
        #informations sur commande

        #Couleurs ratios
        sale_order.ratio_fond_roulement = self.ratio_fond_roulement
        sale_order.ratio_rotation_stock = self.ratio_rotation_stock
        sale_order.ratio_besoin_fond_roulement = self.ratio_besoin_fond_roulement
        sale_order.ratio_clients = self.ratio_clients
        sale_order.ratio_treso_net = self.ratio_treso_net
        sale_order.ratio_rotation_fournisseurs = self.ratio_rotation_fournisseurs
        sale_order.ratio_liquid_test = self.ratio_liquid_test
        sale_order.ratio_degree_endettement = self.ratio_degree_endettement
        sale_order.ratio_autonomie_financiere = self.ratio_autonomie_financiere
        sale_order.ratio_dettes_fin_cafg = self.ratio_dettes_fin_cafg
        sale_order.ratio_rn_caff_ht = self.ratio_rn_caff_ht
        sale_order.ratio_cont_exp = self.ratio_cont_exp
        sale_order.ratio_capacite_fin = self.ratio_capacite_fin
        #Couleurs ratios


        #Sale report generation
        companies_string = ''
        for company_id in self._context['active_ids']:
            companies_string+=str(company_id)+','
        sale_order.companies_string = companies_string[:-1]
        sale_order.company_fiscalyear_id1=self.company_fiscalyear_id1.id
        sale_order.company_fiscalyear_id2=self.company_fiscalyear_id2.id
        if self.product_id.is_report and self.product_id.report_id:
            sale_order.sale_order_print_auto(self.product_id.report_id.report_name)
            #self._print_excel(self.product_id.report_id.report_name,'sale.order',sale_order.id)

        #Reporting history creation
        self.env['sphere.reporting.history'].create({'date':fields.datetime.now(),
                                                     'order_id':sale_order.id,
                                                     'partner_id':self.partner_id.id,
                                                     'product_id':self.product_id.id,
                                                     'type':self.action
                                                     })
        #Sale order redirection
        return  {
            'domain': str([('id', 'in', [sale_order.id])]),
            'view_type': 'form',
            'view_mode': 'tree,form',
            'res_model': 'sale.order',
            'view_id': False,
            'type': 'ir.actions.act_window',
            'name' : _('Devis'),
            'res_id': sale_order.id
                    }

report_order_wizard()